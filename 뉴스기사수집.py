# -*- coding: utf-8 -*-
"""
네이버 뉴스 수집 v6 + Streamlit UI
- 기존 v6 로직 그대로 사용 (requests+BS4 전용)
- 웹 UI에서:
  * 고객사 엑셀(A열) 기반 수집 ("값" +사고)  → 시트명: 고객사 (A=검색어, B=제목, C=기사내용, D=링크)
  * 사용자 지정 검색어(줄 단위)              → 시트명: 사용자 지정 (A=검색어, B=제목, C=기사내용, D=링크)
  * 전력시장 +에너지                          → 시트명: 전력시장 동향 (A=제목, B=기사내용, C=링크)  ※검색어 컬럼 없음
  * N 최대 1~10(유니크)
"""

import os, re, io, time, random, logging, sys
from datetime import datetime
from urllib.parse import quote, urlparse, urlunparse, parse_qs, urlencode

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import streamlit as st

# ===== 기본 설정 =====
BASE_URL = "https://search.naver.com/search.naver?ssc=tab.news.all&where=news&sm=tab_jum&query={query}"
HOMEPAGE = "https://www.naver.com/"
NAVER_QUERY_DELAY_RANGE = (0.8, 1.6)
RETRY_BACKOFF_BASE = 1.5
MAX_RETRY_DEFAULT = 2

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
]
COMMON_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.6,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Referer": HOMEPAGE,
}

# 시간: <span class="sds-comps-text sds-comps-text-type-body2 sds-comps-text-weight-sm">1시간 전</span>
TIME_SPAN_CLASS = ["sds-comps-text", "sds-comps-text-type-body2", "sds-comps-text-weight-sm"]
SNIPPET_CLASSES = ["sds-comps-text-ellipsis", "sds-comps-text-type-body2"]
TRACKING_PARAMS = {"utm_source","utm_medium","utm_campaign","utm_term","utm_content","utm_name",
                   "gclid","fbclid","igshid","utm_id","utm_referrer","ref","sns","spm","cmpid"}

# ===== 로거(간단) =====
logger = logging.getLogger("naver_news_streamlit")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%H:%M:%S"))
    logger.addHandler(sh)

# ===== 유틸 =====
def jitter_sleep(a=NAVER_QUERY_DELAY_RANGE[0], b=NAVER_QUERY_DELAY_RANGE[1]):
    time.sleep(random.uniform(a, b))

def sanitize_query(q: str) -> str:
    # 따옴표는 보존 (앞서 문제였던 strip('"“”') 제거)
    return q.strip()

def build_url(query: str) -> str:
    q = sanitize_query(query)
    return BASE_URL.format(query=quote(q))

def parse_relative_allowed(text: str) -> bool:
    t = text.strip()
    return bool(
        re.fullmatch(r"([1-9]|[1-5][0-9])\s*분\s*전", t)
        or re.fullmatch(r"([1-9]|1[0-9]|2[0-3])\s*시간\s*전", t)
    )

def has_classes(tag, classes) -> bool:
    return bool(tag and tag.has_attr("class") and all(c in tag["class"] for c in classes))

def normalize_title(t: str) -> str:
    return " ".join((t or "").casefold().split())

def normalize_link(url: str) -> str:
    if not url: return ""
    try:
        p = urlparse(url)
        scheme = p.scheme.lower() or "https"
        netloc = p.netloc.lower()
        if netloc.endswith("news.naver.com"):
            qs = parse_qs(p.query)
            oid = qs.get("oid", [None])[0]
            aid = qs.get("aid", [None])[0]
            if oid and aid:
                return f"naver:oid={oid}&aid={aid}"
            return f"{netloc}{p.path.rstrip('/')}"
        qs = parse_qs(p.query, keep_blank_values=True)
        qs_clean = {k: v for k, v in qs.items() if k not in TRACKING_PARAMS}
        qs_clean = {k: sorted(v) for k, v in qs_clean.items()}
        new_query = urlencode(qs_clean, doseq=True)
        new_path = p.path.rstrip("/") or "/"
        return urlunparse((scheme, netloc, new_path, "", new_query, ""))
    except Exception:
        return url

# ===== 네트워크 =====
def make_session(max_retry=MAX_RETRY_DEFAULT) -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": random.choice(UA_POOL), **COMMON_HEADERS})
    try:
        r = s.get(HOMEPAGE, timeout=8)
        r.raise_for_status()
        logger.info("네이버 쿠키 워밍업 성공")
    except Exception as e:
        logger.info(f"워밍업 실패: {e}")
    s._max_retry = max_retry
    return s

def get_html(session: requests.Session, url: str) -> str | None:
    for attempt in range(1, getattr(session, "_max_retry", MAX_RETRY_DEFAULT) + 1):
        try:
            r = session.get(url, timeout=15, allow_redirects=True)
            if r.status_code == 200:
                return r.text
            logger.info(f"HTTP {r.status_code} (시도 {attempt})")
        except Exception as e:
            logger.info(f"요청 실패 (시도 {attempt}): {e}")
        time.sleep(RETRY_BACKOFF_BASE ** attempt)
    return None

# ===== 파서 =====
def extract_card_from_time_span(span):
    a = None; steps = 0
    for e in span.next_elements:
        steps += 1
        if steps > 200: break
        if isinstance(e, str): continue
        if getattr(e, "name", None) == "a" and e.get("data-heatmap-target") == ".tit":
            a = e; break
    if not a: return None
    title = a.get_text(strip=True)
    link  = a.get("href", "")

    snippet = ""
    steps = 0
    for e in a.next_elements:
        steps += 1
        if steps > 200: break
        if getattr(e, "name", None) == "span" and e.has_attr("class"):
            if any(c in e["class"] for c in SNIPPET_CLASSES):
                txt = e.get_text(" ", strip=True)
                if txt and txt != title and len(txt) >= 10:
                    snippet = txt; break
    return {"title": title, "link": link, "snippet": snippet}

# ===== 수집 =====
def fetch_news(session: requests.Session, query: str, max_n: int, include_query_col: bool) -> list[dict]:
    url = build_url(query)
    logger.info(f"검색: {query}")
    html = get_html(session, url)
    if not html:
        logger.info("HTML 획득 실패")
        return []

    soup = BeautifulSoup(html, "html.parser")
    spans = [s for s in soup.find_all("span", class_=lambda x: x) if has_classes(s, TIME_SPAN_CLASS)]

    seen_titles, seen_links, rows = set(), set(), []
    for s in spans:
        if len(rows) >= max_n: break
        t = s.get_text(strip=True)
        if not parse_relative_allowed(t): continue
        card = extract_card_from_time_span(s)
        if not card: continue

        link_norm  = normalize_link(card["link"])
        title_norm = normalize_title(card["title"])
        if (link_norm and link_norm in seen_links) or (title_norm in seen_titles):
            continue
        if link_norm: seen_links.add(link_norm)
        seen_titles.add(title_norm)

        row = {"title": card["title"], "snippet": card["snippet"], "link": card["link"]}
        if include_query_col: row["query"] = query
        rows.append(row)
    jitter_sleep()
    return rows

# ===== 엑셀 헬퍼 =====
def write_sheet(ws, rows: list[dict], include_query_col: bool):
    if include_query_col:
        ws.append(["검색어", "텍스트(제목)", "기사내용", "링크"])
        widths = [28, 60, 100, 80]
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
        for r in rows: ws.append([r.get("query",""), r.get("title",""), r.get("snippet",""), r.get("link","")])
    else:
        ws.append(["텍스트(제목)", "기사내용", "링크"])
        widths = [60, 100, 80]
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
        for r in rows: ws.append([r.get("title",""), r.get("snippet",""), r.get("link","")])

def build_workbook(data_clients: list[dict], data_custom: list[dict] | None, data_market: list[dict]):
    wb = Workbook()

    # 시트1: 고객사
    ws1 = wb.active
    ws1.title = "고객사"
    write_sheet(ws1, data_clients, include_query_col=True)

    # 시트2: 사용자 지정 (있을 때만)
    if data_custom is not None:
        ws2 = wb.create_sheet("사용자 지정")
        write_sheet(ws2, data_custom, include_query_col=True)

    # 시트3: 전력시장 동향
    ws3 = wb.create_sheet("전력시장 동향")
    write_sheet(ws3, data_market, include_query_col=False)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ===== Streamlit UI =====
st.set_page_config(page_title="네이버 뉴스 수집기", layout="wide")
st.title("📰 네이버 뉴스 수집기 (v6 + Streamlit)")

with st.sidebar:
    st.header("설정")
    max_n = st.slider("최대 수집 수 (유니크)", 1, 10, 10)
    custom_mode = st.checkbox("사용자 지정 검색어 모드", value=False, help='텍스트박스에 줄 단위로 검색어를 직접 입력합니다.')
    include_market = st.checkbox("전력시장 +에너지 수집", value=True)
    uploaded = st.file_uploader("고객사 엑셀 업로드 (A열)", type=["xlsx"])
    st.caption("※ 고객사/사용자 지정은 둘 다 체크 시 **합쳐서** 수집합니다.")
    run_btn = st.button("검색 시작")

# 입력 영역
custom_queries = []
if custom_mode:
    st.subheader("사용자 지정 검색어 입력")
    seed = st.text_area("검색어를 줄 단위로 입력하세요. (예: \"홈플러스\" +사고)", height=140)
    if seed.strip():
        custom_queries = [line.strip() for line in seed.splitlines() if line.strip()]

# 실행
if run_btn:
    session = make_session()
    clients_rows_all: list[dict] = []
    custom_rows_all: list[dict] | None = [] if custom_mode else None
    market_rows_all: list[dict] = []

    # 고객사 쿼리
    client_queries = []
    if uploaded is not None:
        try:
            df = pd.read_excel(uploaded, header=None)
            col = df.iloc[:, 0].dropna().astype(str).str.strip()
            # "값" +사고 형태로 생성 (따옴표 유지)
            client_queries = [f'"{v}" +사고' for v in col if v]
        except Exception as e:
            st.error(f"엑셀 읽기 실패: {e}")

    # 사용자 지정 쿼리
    if custom_mode and custom_queries:
        # 그대로 사용 (예: 이미 " +사고" 포함한 상태로 입력했다고 가정)
        pass

    # 고객사 실행
    if client_queries:
        st.info(f"고객사 {len(client_queries)}건 수집 중…")
        progress = st.progress(0.0)
        for idx, q in enumerate(client_queries, start=1):
            rows = fetch_news(session, q, max_n=max_n, include_query_col=True)
            clients_rows_all.extend(rows)
            progress.progress(idx / max(1, len(client_queries)))
        st.success(f"고객사 수집 완료: {len(clients_rows_all)}건")

    # 사용자 지정 실행
    if custom_mode and custom_queries:
        st.info(f"사용자 지정 {len(custom_queries)}건 수집 중…")
        progress = st.progress(0.0)
        for idx, q in enumerate(custom_queries, start=1):
            rows = fetch_news(session, q, max_n=max_n, include_query_col=True)
            custom_rows_all.extend(rows)
            progress.progress(idx / max(1, len(custom_queries)))
        st.success(f"사용자 지정 수집 완료: {len(custom_rows_all)}건")

    # 전력시장 동향
    if include_market:
        st.info("전력시장 동향 수집 중…")
        market_rows_all = fetch_news(session, "전력시장 +에너지", max_n=max_n, include_query_col=False)
        st.success(f"전력시장 동향 수집 완료: {len(market_rows_all)}건")

    # 표시 & 다운로드
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("고객사 결과 미리보기")
        df_clients = pd.DataFrame(clients_rows_all) if clients_rows_all else pd.DataFrame(columns=["query","title","snippet","link"])
        st.dataframe(df_clients)

    with col2:
        st.subheader("전력시장 동향 미리보기")
        df_market = pd.DataFrame(market_rows_all) if market_rows_all else pd.DataFrame(columns=["title","snippet","link"])
        st.dataframe(df_market)

    if custom_mode:
        st.subheader("사용자 지정 결과 미리보기")
        df_custom = pd.DataFrame(custom_rows_all) if custom_rows_all else pd.DataFrame(columns=["query","title","snippet","link"])
        st.dataframe(df_custom)

    # 엑셀 다운로드
    bio = build_workbook(
        data_clients=clients_rows_all,
        data_custom=custom_rows_all if custom_mode else None,
        data_market=market_rows_all
    )
    out_name = f"기사수집_{datetime.now().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="📥 엑셀 다운로드",
        data=bio.getvalue(),
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



