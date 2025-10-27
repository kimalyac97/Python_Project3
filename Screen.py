# -*- coding: utf-8 -*-
"""
pp.kepco.co.kr 배치 자동화 – Streamlit 포팅판 (기능 동일)
- 핵심 동작(로그인/세션 리셋/고객정보 진입/탭 열기/3필드 수집/엑셀 즉시 기록/디버그 저장)은 원본과 동일
- 차이점: CLI→Streamlit UI, Windows MessageBox→화면 알림, tqdm→화면 진행바
- 셀레니움/크롬드라이버는 원본과 동일하게 사용

실행 방법:
$ streamlit run streamlit_app.py

주의:
- 서버 환경(예: 리눅스/원격)에서는 Chrome & chromedriver 설치 필요
- 사이트 구조 변경/차단 정책 등 환경 요인에 따라 동작이 달라질 수 있음
"""

import os, re, sys, time, zipfile
from io import BytesIO
from pathlib import Path
from typing import Tuple, Optional
from datetime import datetime

import pandas as pd
from dotenv import load_dotenv

import streamlit as st

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
    UnexpectedAlertPresentException,
)

from openpyxl import Workbook, load_workbook

# ===== 상수/경로 (기본값; UI에서 변경/주입) =====
URL_BASE = "https://pp.kepco.co.kr"
URL_INTRO = f"{URL_BASE}/intro.do"
CUSTOMER_INFO_PATH = "/mb/mb0101.do?menu_id=O010601"

HEADERS = ["자원명", "고객사명", "ID(고객번호)", "PW", "계기번호", "한전 계약전력(kW)", "계약종별"]

# 동적 설정(초기값) — UI로 갱신
ENABLE_DEBUG_DUMP = True
HEADLESS = True
SCREEN_BASE = Path.cwd() / "screenshots"
DEBUG_BASE = Path.cwd() / "debug"

# ===== 유틸 =====
def sanitize_sheet(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", str(name).strip() or "NONAME")[:31]

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(name).strip() or "NONAME")

# ===== 드라이버 =====
def build_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    )
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=ko-KR")
    if HEADLESS:
        # Streamlit 서버 환경 고려: new headless
        opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1920,1080")
    opts.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    opts.add_experimental_option('useAutomationExtension', False)
    opts.add_argument("--disable-logging")
    opts.add_argument("--log-level=3")

    # chromedriver 경로를 명시해야 하는 환경이라면 Service(executable_path=...) 지정
    service = Service(log_path=os.devnull)
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(60)
    return driver


def reset_session(driver: Optional[webdriver.Chrome]) -> webdriver.Chrome:
    try:
        if driver:
            driver.quit()
    except Exception:
        pass
    time.sleep(0.5)
    return build_driver()


def wait_ready(driver, timeout=20):
    WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")


def wait_click(driver, by, value, timeout=15):
    elem = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
    try:
        elem.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", elem)
    return elem


def wait_sendkeys(driver, by, value, text, timeout=15):
    elem = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
    try:
        elem.clear()
    except Exception:
        pass
    elem.send_keys(text)
    return elem

# ===== 로그인 성공 판정 =====
def _is_logged_in(driver, timeout=20) -> bool:
    end = time.time() + timeout
    while time.time() < end:
        if "intro.do" not in (driver.current_url or "").lower():
            return True
        time.sleep(0.5)
    return False


def run_once_with_credentials(driver, user_id: str, user_pw: str) -> Tuple[bool, Optional[str]]:
    try:
        driver.get(URL_INTRO)
        wait_ready(driver, 20)
        wait_sendkeys(driver, By.ID, "RSA_USER_ID", user_id)
        wait_sendkeys(driver, By.ID, "RSA_USER_PWD", user_pw)
        wait_click(driver, By.CSS_SELECTOR, 'input.intro_btn[type="button"][value="로그인"]')
        if _is_logged_in(driver, timeout=25):
            return True, None
        return False, "LOGIN_FAILED"
    except UnexpectedAlertPresentException:
        try:
            driver.switch_to.alert.accept()
        except Exception:
            pass
        return False, "UNEXPECTED_ALERT"
    except Exception as e:
        return False, repr(e)

# ===== 고객정보 진입/탭 열기/필드 수집 =====
def goto_customer_info(driver):
    driver.get(URL_BASE + CUSTOMER_INFO_PATH)
    wait_ready(driver, 20)
    time.sleep(0.6)
    return True


def open_meter_tab(driver, timeout=10) -> bool:
    try:
        link = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, \"tabTable('3')\")]"))
        )
        driver.execute_script("arguments[0].click();", link)
    except TimeoutException:
        try:
            driver.execute_script("tabTable('3')")
        except Exception:
            return False

    end = time.time() + timeout
    while time.time() < end:
        try:
            if driver.find_elements(By.CSS_SELECTOR, "#table3"):
                return True
        except Exception:
            pass
        time.sleep(0.3)
    return True


def fetch_three_fields(driver) -> Tuple[str, str, str, Optional[str]]:
    try:
        meter = driver.find_element(By.CSS_SELECTOR, "#table3 tbody tr:nth-child(1) td:nth-child(2)").text.strip()
        kw    = driver.find_element(By.CSS_SELECTOR, "div.table_info table tbody tr:nth-child(2) td:nth-child(4)").text.strip()
        ctype = driver.find_element(By.CSS_SELECTOR, "div.table_info table tbody tr:nth-child(2) td:nth-child(2)").text.strip()
        return meter, kw, ctype, None
    except Exception:
        return "", "", "", "FIELD_NOT_FOUND"

# ===== 스크린샷/디버그 =====
def center_mouse_and_screenshot(driver, sheet_name: str, cust_name: str):
    try:
        fname = sanitize_filename(cust_name) + ".png"
        save_path = SCREEN_BASE / sanitize_sheet(sheet_name) / fname
        save_path.parent.mkdir(parents=True, exist_ok=True)
        driver.save_screenshot(str(save_path))
        return True, str(save_path)
    except Exception as e:
        return False, repr(e)


def dump_debug_html(driver, sheet_name: str, cust_name: str):
    if not ENABLE_DEBUG_DUMP:
        return
    try:
        fname = sanitize_filename(cust_name) + "_page.html"
        path = DEBUG_BASE / sanitize_sheet(sheet_name) / fname
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(driver.page_source, encoding="utf-8", errors="ignore")
    except Exception:
        pass

# ===== 엑셀 유틸 =====
def ensure_workbook(path: Path):
    if not path.exists():
        Workbook().save(path)


def ensure_sheet_with_header(path: Path, sheet_name: str):
    sheet_name = sanitize_sheet(sheet_name)
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(HEADERS)
    wb.save(path)
    wb.close()
    return sheet_name


def append_row(path: Path, sheet_name: str, row_values: list):
    sheet_name = sanitize_sheet(sheet_name)
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.append(row_values)
    wb.save(path)
    wb.close()

# ===== 계정 처리(원본 로직 유지) =====
def process_account(sheet_name: str, cust_name: str, user_id: str, user_pw: str) -> Tuple[str, str, str, bool]:
    driver = reset_session(None)   # 계정마다 완전 새 세션 시작
    ok, reason = run_once_with_credentials(driver, user_id, user_pw)
    if not ok:
        driver.quit()
        return "", "", "", False

    try:
        goto_customer_info(driver)
        center_mouse_and_screenshot(driver, sheet_name, cust_name)

        if not open_meter_tab(driver, timeout=10):
            dump_debug_html(driver, sheet_name, cust_name)
            driver.quit()
            return "", "", "", False

        meter, kw, ctype, err = fetch_three_fields(driver)
        driver.quit()
        if err:
            dump_debug_html(driver, sheet_name, cust_name)
            return "", "", "", False
        return meter, kw, ctype, True
    except Exception:
        dump_debug_html(driver, sheet_name, cust_name)
        driver.quit()
        return "", "", "", False

# ===== 배치 실행(Progress 콜백 반영) =====
def read_excel_all_sheets(xlsx_path: Path) -> dict:
    xls = pd.ExcelFile(xlsx_path)
    return {s: pd.read_excel(xlsx_path, sheet_name=s, dtype=str).fillna("").astype(str) for s in xls.sheet_names}


def run_batch(excel_path: Path, out_xlsx_name: Path, progress_cb=None, log_cb=None):
    load_dotenv()
    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 없음: {excel_path.resolve()}")

    # 출력 폴더 준비
    SCREEN_BASE.mkdir(parents=True, exist_ok=True)
    DEBUG_BASE.mkdir(parents=True, exist_ok=True)

    sheets = read_excel_all_sheets(excel_path)
    total_jobs = sum(
        len(df[(df.get("ID", "") != "") & (df.get("PW", "") != "")])
        for df in sheets.values()
        if "ID" in df and "PW" in df and "고객명" in df
    )

    ensure_workbook(out_xlsx_name)
    for s in sheets.keys():
        ensure_sheet_with_header(out_xlsx_name, s)

    success_cnt, fail_cnt = 0, 0
    done = 0

    for sheet_name, df in sheets.items():
        if not all(c in df.columns for c in ("ID", "PW", "고객명")):
            continue
        safe_sheet = ensure_sheet_with_header(out_xlsx_name, sheet_name)
        for _, row in df.iterrows():
            user_id = str(row["ID"]).strip()
            user_pw = str(row["PW"]).strip()
            cust_name = str(row["고객명"]).strip()
            if not user_id or not user_pw:
                continue

            if log_cb:
                log_cb(f"처리 중: [{sheet_name}] {cust_name} ({user_id})")

            meter, kw, ctype, ok = process_account(sheet_name, cust_name, user_id, user_pw)
            if ok:
                success_cnt += 1
            else:
                fail_cnt += 1
            append_row(out_xlsx_name, safe_sheet, [sheet_name, cust_name, user_id, user_pw, meter, kw, ctype])

            done += 1
            if progress_cb and total_jobs:
                progress_cb(done / total_jobs)

    summary = f"총 작업 {total_jobs}건\n성공 {success_cnt}건, 실패 {fail_cnt}건\n결과: {out_xlsx_name.resolve()}"
    return summary, success_cnt, fail_cnt, total_jobs


# ===== Streamlit UI =====
st.set_page_config(page_title="파워플래너 정보 취합 (KEPCO)", layout="wide")
st.title("파워플래너 정보 취합 – Streamlit")

with st.sidebar:
    st.header("실행 옵션")
    HEADLESS = st.checkbox("Headless(권장)", value=True)
    ENABLE_DEBUG_DUMP = st.checkbox("디버그 HTML 저장", value=True)

    st.caption("스크린샷/디버그 저장 위치는 현재 작업 폴더의 'screenshots', 'debug'입니다.")

    up = st.file_uploader("입력 엑셀 업로드 (원본: 20251010고객사id,pw2.xlsx)", type=["xlsx"])
    excel_path: Optional[Path] = None
    if up is not None:
        # 업로드 파일을 임시 경로에 저장
        tmp_path = Path(st.session_state.get("_uploaded_excel_path", f"uploaded_{int(time.time())}.xlsx"))
        with open(tmp_path, "wb") as f:
            f.write(up.getbuffer())
        st.session_state["_uploaded_excel_path"] = str(tmp_path)
        excel_path = tmp_path
        st.success(f"업로드됨: {tmp_path}")
    else:
        # 로컬 경로 직접 입력 옵션(선택)
        default_name = "20251010고객사id,pw2.xlsx"
        manual = st.text_input("또는 로컬 경로 직접 입력", value=default_name)
        if manual:
            excel_path = Path(manual)

# 출력 파일명(원본 규칙 유지)
out_name_default = f"(자원등록용)파워플래너 정보 취합 초안_{datetime.now().strftime('%Y%m%d')}.xlsx"
out_name = st.text_input("출력 엑셀 파일명", value=out_name_default)
out_path = Path(out_name)

col1, col2 = st.columns([1, 2])
run_btn = col1.button("실행")
log_box = st.empty()
progress = st.progress(0.0)
result_box = st.empty()


def _log(msg: str):
    # 누적 로그 출력
    prev = st.session_state.get("_logs", "")
    new = prev + (msg + "\n")
    st.session_state["_logs"] = new
    log_box.text(new)


if run_btn:
    if not excel_path:
        st.error("엑셀 파일을 업로드하거나 경로를 입력하세요.")
        st.stop()

    # 화면 초기화
    st.session_state["_logs"] = ""
    progress.progress(0.0)
    # 전역 경로(스크린샷/디버그) 재지정 가능 — 기본은 CWD
    SCREEN_BASE = Path.cwd() / "screenshots"
    DEBUG_BASE = Path.cwd() / "debug"

    # 배치 실행
    try:
        summary, ok_cnt, fail_cnt, total = run_batch(
            excel_path=excel_path,
            out_xlsx_name=out_path,
            progress_cb=lambda v: progress.progress(min(max(v, 0.0), 1.0)),
            log_cb=_log,
        )
        st.success("작업 완료")
        result_box.code(summary, language="text")

        # 결과 파일 다운로드 버튼
        if out_path.exists():
            with open(out_path, "rb") as f:
                st.download_button(
                    label="결과 엑셀 다운로드",
                    data=f.read(),
                    file_name=out_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        # 스크린샷/디버그 ZIP 다운로드(선택)
        with st.expander("스크린샷/디버그 압축 다운로드"):
            colz1, colz2 = st.columns(2)
            if colz1.button("screenshots 폴더 ZIP 생성"):
                buf = BytesIO()
                with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    base = SCREEN_BASE
                    if base.exists():
                        for p in base.rglob('*'):
                            if p.is_file():
                                zf.write(p, p.relative_to(base))
                buf.seek(0)
                st.download_button("screenshots.zip 다운로드", data=buf.read(), file_name="screenshots.zip", mime="application/zip")

            if colz2.button("debug 폴더 ZIP 생성"):
                buf = BytesIO()
                with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    base = DEBUG_BASE
                    if base.exists():
                        for p in base.rglob('*'):
                            if p.is_file():
                                zf.write(p, p.relative_to(base))
                buf.seek(0)
                st.download_button("debug.zip 다운로드", data=buf.read(), file_name="debug.zip", mime="application/zip")

    except Exception as e:
        st.exception(e)
